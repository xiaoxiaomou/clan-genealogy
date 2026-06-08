from flask import Blueprint, jsonify, make_response, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.services.family_service import FamilyService
from app.utils.decorators import family_permission_required, admin_required
from app.utils.export_utils import generate_pdf, generate_traditional_pdf
from app import db
import io
import csv
import re

export_bp = Blueprint('export', __name__)


@export_bp.route('/<int:family_id>/export/pdf', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_family_pdf(family_id):
    """导出族谱为 PDF"""
    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    members = FamilyService.get_family_members(family_id)
    relationships = FamilyService.get_family_relationships(family_id)
    tree_data = FamilyService.get_family_tree(family_id)

    pdf_data = generate_pdf(family, members, relationships, tree_data)

    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={family.name}.pdf'
    return response


@export_bp.route('/<int:family_id>/export/pdf/traditional', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_traditional_pdf(family_id):
    """导出传统谱式 PDF（苏式/欧式/宝塔式）"""
    style = request.args.get('style', 'su')

    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    members = FamilyService.get_family_members(family_id)
    relationships = FamilyService.get_family_relationships(family_id)
    tree_data = FamilyService.get_family_tree(family_id)

    style_names = {'su': '苏式', 'ou': '欧式', 'bao': '宝塔式'}
    pdf_data = generate_traditional_pdf(family, members, relationships, tree_data, style)

    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f"attachment; filename=family_{family_id}_{style}.pdf"
    return response


@export_bp.route('/<int:family_id>/export/pdf/guzhi', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_guzhi_pdf(family_id):
    """导出古籍线装风格 PDF：宣纸色 + 红色印章 + 仿古标题 + 跋文"""
    from app.utils.export_utils import generate_guzhi_pdf

    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    members = FamilyService.get_family_members(family_id)
    relationships = FamilyService.get_family_relationships(family_id)
    tree_data = FamilyService.get_family_tree(family_id)

    try:
        pdf_data = generate_guzhi_pdf(family, members, relationships, tree_data)
    except Exception as e:
        return jsonify({'error': f'生成 PDF 失败：{str(e)}'}), 500

    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f"attachment; filename=family_{family_id}_guzhi.pdf"
    return response


@export_bp.route('/<int:family_id>/export/members/csv', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_members_csv(family_id):
    """导出族谱成员为 CSV"""
    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    members = FamilyService.get_family_members(family_id)

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ['name', 'gender', 'birth_date', 'death_date', 'generation', 'generation_name', 'bio', 'is_alive']
    header_labels = ['姓名', '性别', '出生日期', '逝世日期', '辈分', '字辈', '简介', '是否在世']
    writer.writerow(header_labels)

    for m in members:
        writer.writerow([
            m.name,
            m.gender,
            m.birth_date or '',
            m.death_date or '',
            m.generation or '',
            m.generation_name or '',
            m.bio or '',
            1 if m.is_alive else 0,
        ])

    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv;charset=utf-8-sig'
    response.headers['Content-Disposition'] = f'attachment; filename={family.name}_members.csv'
    return response


@export_bp.route('/<int:family_id>/export/members/excel', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_members_excel(family_id):
    """导出族谱成员为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    members = FamilyService.get_family_members(family_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '家族成员'

    headers = ['姓名', '性别', '出生日期', '逝世日期', '辈分', '字辈', '简介', '是否在世']
    header_fill = PatternFill(start_color='8b2500', end_color='8b2500', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    gender_map = {'male': '男', 'female': '女', 'unknown': '未知'}
    for row_idx, m in enumerate(members, start=2):
        ws.cell(row=row_idx, column=1, value=m.name).border = thin_border
        ws.cell(row=row_idx, column=2, value=gender_map.get(m.gender, '未知')).border = thin_border
        ws.cell(row=row_idx, column=3, value=m.birth_date or '').border = thin_border
        ws.cell(row=row_idx, column=4, value=m.death_date or '').border = thin_border
        ws.cell(row=row_idx, column=5, value=m.generation or '').border = thin_border
        ws.cell(row=row_idx, column=6, value=m.generation_name or '').border = thin_border
        ws.cell(row=row_idx, column=7, value=m.bio or '').border = thin_border
        ws.cell(row=row_idx, column=8, value='在世' if m.is_alive else '已故').border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={family.name}_members.xlsx'
    return response


@export_bp.route('/<int:family_id>/export/gedcom', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def export_gedcom(family_id):
    """导出族谱为 GEDCOM 5.5.1 格式"""
    from app.services.gedcom_service import export_family

    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    try:
        gedcom_text = export_family(family_id)
    except ValueError as e:
        return jsonify({'error': str(e)}), 404

    response = make_response(gedcom_text)
    response.headers['Content-Type'] = 'text/plain;charset=utf-8'
    # 全部 ASCII 文件名，避免 werkzeug latin-1 编码错误
    response.headers['Content-Disposition'] = f"attachment; filename=family_{family_id}.ged"
    return response


@export_bp.route('/<int:family_id>/import/gedcom', methods=['POST'])
@jwt_required()
@admin_required
@family_permission_required('editor')
def import_gedcom(family_id):
    """从 GEDCOM 格式导入族谱"""
    from app.services.gedcom_service import import_family as ged_import

    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404

    if 'file' not in request.files:
        return jsonify({'error': '未找到上传文件'}), 400

    file = request.files['file']
    if not (file.filename or '').lower().endswith('.ged'):
        return jsonify({'error': '请上传 .ged 文件'}), 400

    raw = file.read()
    try:
        gedcom_text = raw.decode('utf-8')
    except UnicodeDecodeError:
        gedcom_text = raw.decode('gb18030', errors='ignore')

    try:
        result = ged_import(family_id, gedcom_text)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    return jsonify({
        'message': f"成功导入 {result['imported']} 个成员，{result['relations']} 条关系",
        'members_count': result['imported'],
        'relations_count': result['relations'],
    })