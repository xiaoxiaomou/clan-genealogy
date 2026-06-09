from flask import Blueprint, request, jsonify, g
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.services.family_service import FamilyService
from app.utils.decorators import family_permission_required, admin_required
from app.utils.html_utils import sanitize_html
from app.models.user import User
from app.models.member import Member
from app.models.member_edit_history import MemberEditHistory
from app.routes.member_history import record_member_changes
from app.services.privacy_service import filter_member_list, filter_member_dict

family_bp = Blueprint('family', __name__)


def sanitize_bio(bio):
    """对成员 bio 字段做 HTML 安全清理，富文本支持"""
    if not bio:
        return bio
    return sanitize_html(bio, max_length=20000)


# ==================== 族谱管理 ====================

@family_bp.route('/', methods=['GET'])
@jwt_required()
def list_families():
    """获取用户的所有族谱"""
    user_id = int(get_jwt_identity())
    families = FamilyService.get_user_families(user_id)
    return jsonify({
        'families': [f.to_dict() for f in families]
    })


@family_bp.route('/<int:family_id>/my-role', methods=['GET'])
@jwt_required()
def get_my_role(family_id):
    """获取当前用户在族谱中的角色"""
    user_id = int(get_jwt_identity())

    # 管理员自动拥有所有权限
    user = db.session.get(User, user_id)
    if user and user.is_admin:
        return jsonify({'role': 'admin', 'can_edit': True})

    fm = FamilyService.get_family_users(family_id)
    my_record = next((u for u in fm if u['user_id'] == user_id), None)
    if my_record:
        role = my_record['role']
        can_edit = role in ('owner', 'admin', 'editor')
        return jsonify({'role': role, 'can_edit': can_edit})

    # 不在族谱中，检查是否公开
    from app.models.family import Family
    family = db.session.get(Family, family_id)
    if family and family.is_public:
        return jsonify({'role': 'viewer', 'can_edit': False})

    return jsonify({'error': '无权访问该族谱'}), 403


@family_bp.route('/', methods=['POST'])
@jwt_required()
@admin_required
def create_family():
    """创建族谱（仅管理员）"""
    user_id = int(get_jwt_identity())

    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '族谱名称不能为空'}), 400

    family = FamilyService.create_family(
        name=name,
        creator_id=user_id,
        description=data.get('description'),
        surname=data.get('surname'),
        origin=data.get('origin'),
        is_public=data.get('is_public', False)
    )
    return jsonify({
        'message': '族谱创建成功',
        'family': family.to_dict()
    }), 201


@family_bp.route('/<int:family_id>', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_family(family_id):
    """获取族谱详情"""
    family = FamilyService.get_family(family_id)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404
    return jsonify({'family': family.to_dict()})


@family_bp.route('/<int:family_id>', methods=['PUT'])
@jwt_required()
@family_permission_required('admin')
def update_family(family_id):
    """更新族谱信息"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    family = FamilyService.update_family(family_id, **data)
    if not family:
        return jsonify({'error': '族谱不存在'}), 404
    return jsonify({
        'message': '族谱更新成功',
        'family': family.to_dict()
    })


@family_bp.route('/<int:family_id>', methods=['DELETE'])
@jwt_required()
@family_permission_required('owner')
def delete_family(family_id):
    """删除族谱"""
    if FamilyService.delete_family(family_id):
        return jsonify({'message': '族谱删除成功'})
    return jsonify({'error': '族谱不存在'}), 404


# ==================== 成员管理 ====================

def _get_role(family_id):
    """获取当前用户在族谱中的角色"""
    from flask_jwt_extended import verify_jwt_in_request, get_jwt_identity
    try:
        verify_jwt_in_request(optional=True)
        user_id = int(get_jwt_identity())
    except:
        return 'viewer'
    user = db.session.get(User, user_id)
    if user and user.is_admin:
        return 'admin'
    from app.models.family import FamilyMember
    fm = FamilyMember.query.filter_by(family_id=family_id, user_id=user_id).first()
    return fm.role if fm else 'viewer'


@family_bp.route('/<int:family_id>/members', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def list_members(family_id):
    """获取族谱所有成员"""
    members = FamilyService.get_family_members(family_id)
    role = _get_role(family_id)
    data = [m.to_dict() for m in members]
    data = filter_member_list(members, data, role)
    return jsonify({
        'members': data
    })


@family_bp.route('/<int:family_id>/members', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def add_member(family_id):
    """添加家族成员"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '成员姓名不能为空'}), 400

    member = FamilyService.add_member(
        family_id=family_id,
        name=name,
        gender=data.get('gender', 'unknown'),
        birth_date=data.get('birth_date'),
        death_date=data.get('death_date'),
        generation=data.get('generation'),
        generation_name=data.get('generation_name'),
        bio=sanitize_bio(data.get('bio')),
        avatar=data.get('avatar'),
        is_alive=data.get('is_alive', True)
    )
    # 活人保护自动标记
    from app.services.privacy_service import is_living_protected
    if is_living_protected(member):
        member.privacy_level = 'private'
        db.session.commit()
    return jsonify({
        'message': '成员添加成功',
        'member': member.to_dict()
    }), 201


@family_bp.route('/<int:family_id>/members/<int:member_id>', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_member(family_id, member_id):
    """获取成员详情"""
    member = FamilyService.get_member(member_id)
    if not member or member.family_id != family_id:
        return jsonify({'error': '成员不存在'}), 404
    role = _get_role(family_id)
    data = member.to_dict(include_relations=True)
    data = filter_member_dict(data, role)
    return jsonify({'member': data})


@family_bp.route('/<int:family_id>/members/<int:member_id>', methods=['PUT'])
@jwt_required()
@family_permission_required('editor')
def update_member(family_id, member_id):
    """更新成员信息"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    # bio 字段做 HTML 安全清理（支持富文本）
    if 'bio' in data and data['bio']:
        data['bio'] = sanitize_bio(data['bio'])

    # 先读旧值用于 history
    old_member = Member.query.filter_by(id=member_id, family_id=family_id).first()
    if not old_member:
        return jsonify({'error': '成员不存在'}), 404
    old_attrs = {
        f: getattr(old_member, f)
        for f in ('name', 'gender', 'generation', 'birth_date', 'death_date',
                   'is_alive', 'birth_place', 'death_place', 'generation_name',
                   'father_id', 'mother_id', 'spouse_id', 'branch_id', 'bio',
                   'occupation', 'avatar', 'courtesy_name', 'art_name', 'posthumous_name')
        if hasattr(old_member, f)
    }

    # 活人保护同步 & privacy_level 支持
    if 'privacy_level' in data:
        member.privacy_level = data['privacy_level']
        member.privacy_override = True
    from app.services.privacy_service import is_living_protected
    if 'is_alive' in data or 'birth_date' in data:
        if is_living_protected(member) and not member.privacy_override:
            member.privacy_level = 'private'
        elif not is_living_protected(member) and not member.privacy_override:
            member.privacy_level = 'public'
    
    member = FamilyService.update_member(member_id, **{k: v for k, v in data.items() if k not in ('privacy_level', 'privacy_override')})
    if not member:
        return jsonify({'error': '成员不存在'}), 404

    # 记录版本历史
    try:
        uid = int(get_jwt_identity())
        record_member_changes(
            family_id=family_id,
            member=member,
            old_attrs=old_attrs,
            new_attrs=data,
            editor_id=uid,
        )
        db.session.commit()
    except Exception:
        # history 失败不阻塞主流程
        db.session.rollback()

    return jsonify({
        'message': '成员信息更新成功',
        'member': member.to_dict()
    })


@family_bp.route('/<int:family_id>/members/<int:member_id>', methods=['DELETE'])
@jwt_required()
@family_permission_required('editor')
def delete_member(family_id, member_id):
    """删除成员"""
    if FamilyService.delete_member(member_id):
        return jsonify({'message': '成员删除成功'})
    return jsonify({'error': '成员不存在'}), 404


# ==================== 批量操作 ====================

# 允许批量更新的字段白名单
BATCH_EDITABLE_FIELDS = {
    'gender', 'is_alive', 'branch_id', 'generation_name',
    'birth_place', 'death_place',
}


@family_bp.route('/<int:family_id>/members/batch-edit', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def batch_edit_members(family_id):
    """批量编辑多个成员的共享字段

    请求体: {
      member_ids: [1, 2, 3, ...],
      updates: { gender: "male", branch_id: 2, is_alive: false, ... },
      mode: "set" | "append",  # set=覆盖, append=追加（仅对 generation_name 生效）
    }
    """
    data = request.get_json(silent=True) or {}
    member_ids = data.get('member_ids') or []
    updates = data.get('updates') or {}
    mode = data.get('mode', 'set')

    if not isinstance(member_ids, list) or not member_ids:
        return jsonify({'error': 'member_ids 必须是非空数组'}), 400
    if not isinstance(updates, dict) or not updates:
        return jsonify({'error': 'updates 必须是非空对象'}), 400

    # 字段白名单
    invalid = set(updates.keys()) - BATCH_EDITABLE_FIELDS
    if invalid:
        return jsonify({
            'error': f'不允许批量修改字段: {", ".join(sorted(invalid))}',
            'allowed': sorted(BATCH_EDITABLE_FIELDS),
        }), 400

    # 字段级校验
    clean_updates = {}
    for k, v in updates.items():
        if k == 'gender' and v not in ('male', 'female', 'unknown'):
            return jsonify({'error': 'gender 必须是 male/female/unknown'}), 400
        if k == 'is_alive':
            clean_updates[k] = bool(v) if isinstance(v, bool) else str(v).lower() in ('true', '1', 'yes')
        elif k == 'branch_id':
            try:
                clean_updates[k] = int(v) if v not in (None, '', 'null') else None
            except (TypeError, ValueError):
                return jsonify({'error': 'branch_id 必须是整数'}), 400
        else:
            clean_updates[k] = v if v != '' else None

    try:
        affected = 0
        skipped = []
        for mid in member_ids:
            try:
                mid_int = int(mid)
            except (TypeError, ValueError):
                skipped.append({'id': mid, 'reason': 'ID 无效'})
                continue
            m = Member.query.filter_by(id=mid_int, family_id=family_id).first()
            if not m:
                skipped.append({'id': mid_int, 'reason': '成员不存在或不属于本家族'})
                continue
            for k, v in clean_updates.items():
                if mode == 'append' and k == 'generation_name':
                    cur = m.generation_name or ''
                    new_v = v
                    if new_v and new_v not in cur.split(','):
                        m.generation_name = (cur + ',' + new_v).strip(',') if cur else new_v
                else:
                    setattr(m, k, v)
            affected += 1

        db.session.commit()
        # 审计
        from app.utils.audit import log_action
        log_action(
            family_id=family_id,
            user_id=int(get_jwt_identity()),
            action='batch_update',
            entity_type='member',
            description=f'批量更新 {affected} 个成员，字段: {", ".join(clean_updates.keys())} (mode={mode})',
        )
        return jsonify({
            'message': f'成功更新 {affected} 个成员',
            'affected': affected,
            'skipped': skipped,
            'updates_applied': list(clean_updates.keys()),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'批量更新失败: {str(e)}'}), 500


@family_bp.route('/<int:family_id>/members/batch-delete', methods=['POST'])
@jwt_required()
@family_permission_required('admin')
def batch_delete_members(family_id):
    """批量删除（仅 admin 权限）"""
    data = request.get_json(silent=True) or {}
    member_ids = data.get('member_ids') or []
    if not isinstance(member_ids, list) or not member_ids:
        return jsonify({'error': 'member_ids 必须是非空数组'}), 400

    deleted = 0
    failed = []
    for mid in member_ids:
        try:
            mid_int = int(mid)
        except (TypeError, ValueError):
            failed.append({'id': mid, 'reason': 'ID 无效'})
            continue
        if FamilyService.delete_member(mid_int):
            deleted += 1
        else:
            failed.append({'id': mid_int, 'reason': '成员不存在'})

    db.session.commit()
    from app.utils.audit import log_action
    log_action(
        family_id=family_id,
        user_id=int(get_jwt_identity()),
        action='batch_delete',
        entity_type='member',
        description=f'批量删除 {deleted} 个成员',
    )
    return jsonify({
        'message': f'成功删除 {deleted} 个成员',
        'deleted': deleted,
        'failed': failed,
    })


@family_bp.route('/<int:family_id>/members/batch-sort', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def batch_sort_members(family_id):
    """批量设置兄弟节点排序

    请求体: {
      orders: [ { id: 1, sort_order: 0 }, { id: 2, sort_order: 1 }, ... ]
    }
    """
    data = request.get_json(silent=True) or {}
    orders = data.get('orders') or []
    if not isinstance(orders, list) or not orders:
        return jsonify({'error': 'orders 必须是非空数组'}), 400

    updated = 0
    for item in orders:
        mid = item.get('id')
        sort_val = item.get('sort_order')
        if mid is None or sort_val is None:
            continue
        try:
            mid_int = int(mid)
            sort_int = int(sort_val)
        except (TypeError, ValueError):
            continue
        m = Member.query.filter_by(id=mid_int, family_id=family_id).first()
        if m:
            m.sort_order = sort_int
            updated += 1

    db.session.commit()
    from app.utils.audit import log_action
    log_action(
        family_id=family_id,
        user_id=int(get_jwt_identity()),
        action='batch_sort',
        entity_type='member',
        description=f'批量排序 {updated} 个成员',
    )
    return jsonify({'message': f'成功排序 {updated} 个成员', 'updated': updated})


@family_bp.route('/<int:family_id>/members/import-csv', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def import_members_csv(family_id):
    """从 CSV 文本导入成员（dry-run 模式 + 实际写入两阶段）

    请求体: {
      csv_text: "name,gender,birth_date,death_date,generation_name,birth_place,is_alive\n...",
      dry_run: true,  # 仅校验，不写入
      skip_header: true,
    }

    响应: {
      total: 50,
      valid: 48,
      invalid: 2,
      errors: [{ row: 3, name: "张三", reason: "缺少必填 name 字段" }, ...],
      preview: [...first 5 parsed members...]
    }
    """
    import csv as csv_mod
    from io import StringIO

    data = request.get_json(silent=True) or {}
    csv_text = data.get('csv_text', '').strip()
    dry_run = bool(data.get('dry_run', True))
    skip_header = bool(data.get('skip_header', True))

    if not csv_text:
        return jsonify({'error': 'csv_text 不能为空'}), 400

    reader = csv_mod.DictReader(StringIO(csv_text))
    rows = list(reader)

    valid = []
    errors = []
    for i, row in enumerate(rows, start=2 if skip_header else 1):
        name = (row.get('name') or '').strip()
        if not name:
            errors.append({'row': i, 'name': '', 'reason': '缺少必填 name 字段'})
            continue
        gender = (row.get('gender') or 'unknown').strip().lower()
        if gender not in ('male', 'female', 'unknown'):
            errors.append({'row': i, 'name': name, 'reason': f'gender 非法: {gender}'})
            continue
        gen_name = (row.get('generation_name') or '').strip() or None
        try:
            gen = int(row['generation']) if row.get('generation') else None
        except (TypeError, ValueError):
            gen = None
        valid.append({
            'name': name,
            'gender': gender,
            'birth_date': (row.get('birth_date') or '').strip() or None,
            'death_date': (row.get('death_date') or '').strip() or None,
            'generation_name': gen_name,
            'generation': gen,
            'birth_place': (row.get('birth_place') or '').strip() or None,
            'death_place': (row.get('death_place') or '').strip() or None,
            'is_alive': str(row.get('is_alive', 'true')).lower() in ('true', '1', 'yes', '是', '1'),
            'bio': (row.get('bio') or '').strip() or None,
        })

    if dry_run:
        return jsonify({
            'dry_run': True,
            'total': len(rows),
            'valid': len(valid),
            'invalid': len(errors),
            'errors': errors[:30],
            'preview': valid[:5],
        })

    # 实际写入
    try:
        added = 0
        for m_data in valid:
            m = Member(family_id=family_id, **m_data)
            db.session.add(m)
            added += 1
        db.session.commit()
        from app.utils.audit import log_action
        log_action(
            family_id=family_id,
            user_id=int(get_jwt_identity()),
            action='csv_import',
            entity_type='member',
            description=f'CSV 导入 {added} 个成员',
        )
        return jsonify({
            'dry_run': False,
            'added': added,
            'invalid': len(errors),
            'errors': errors[:30],
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'CSV 导入失败: {str(e)}'}), 500


@family_bp.route('/<int:family_id>/import/excel', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def import_members_excel(family_id):
    """从 Excel (.xlsx) 文件导入成员，支持行级错误报告"""
    dry_run = request.form.get('dry_run', 'true').lower() == 'true'
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': '请上传 .xlsx 文件'}), 400
    if not file.filename.endswith('.xlsx') and not file.filename.endswith('.xls'):
        return jsonify({'error': '仅支持 .xlsx 格式'}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

    if len(rows) < 2:
        return jsonify({'error': '文件为空或只有表头'}), 400

    headers = [str(h or '').strip().lower().replace(' ', '_') for h in rows[0]]
    required_field = 'name'

    COLUMN_MAP = {
        'name': 'name', '姓名': 'name', '名字': 'name',
        'gender': 'gender', '性别': 'gender',
        'birth_date': 'birth_date', '出生日期': 'birth_date', '出生': 'birth_date',
        'death_date': 'death_date', '逝世日期': 'death_date', '逝世': 'death_date',
        'generation': 'generation', '辈分': 'generation', '第几代': 'generation', '世代': 'generation',
        'generation_name': 'generation_name', '辈分名称': 'generation_name', '字辈': 'generation_name',
        'birth_place': 'birth_place', '出生地': 'birth_place',
        'death_place': 'death_place', '逝世地': 'death_place',
        'is_alive': 'is_alive', '是否在世': 'is_alive', '在世': 'is_alive',
        'bio': 'bio', '简介': 'bio', '生平': 'bio',
        'courtesy_name': 'courtesy_name', '字': 'courtesy_name',
        'art_name': 'art_name', '号': 'art_name',
        'posthumous_name': 'posthumous_name', '谥': 'posthumous_name', '谥号': 'posthumous_name',
    }

    field_map = {}
    for i, h in enumerate(headers):
        mapped = COLUMN_MAP.get(h)
        if mapped:
            field_map[i] = mapped

    valid = []
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        name = str(row[0]).strip() if len(row) > 0 and row[0] else ''
        if not name:
            errors.append({'row': row_idx, 'name': '', 'reason': '缺少姓名', 'field': 'name'})
            continue

        member_data = {'name': name, 'family_id': family_id}
        col_errors = []
        for col_idx, value in enumerate(row[1:], start=1):
            field = field_map.get(col_idx)
            if not field:
                continue
            if value is None:
                continue
            val_str = str(value).strip()

            if field == 'gender':
                mapped = {'男': 'male', '女': 'female', '未知': 'unknown', 'male': 'male', 'female': 'female', 'unknown': 'unknown'}
                val_str = mapped.get(val_str.lower(), 'unknown')
            elif field == 'is_alive':
                val_str = str(val_str).lower() in ('true', '1', 'yes', '是', '1', '活着')
            elif field == 'generation':
                try:
                    val_str = int(float(val_str))
                except (ValueError, TypeError):
                    col_errors.append(f'第 {col_idx+1} 列「{field}」无法解析为数字')
                    continue

            member_data[field] = val_str

        if col_errors:
            errors.append({'row': row_idx, 'name': name, 'reason': '; '.join(col_errors), 'field': 'multiple'})
            continue

        valid.append(member_data)

    if dry_run:
        return jsonify({
            'dry_run': True,
            'total': len(rows) - 1,
            'valid': len(valid),
            'invalid': len(errors),
            'errors': errors[:50],
            'preview': valid[:5],
        })

    try:
        added = 0
        for m_data in valid:
            m = Member(**m_data)
            db.session.add(m)
            added += 1
        db.session.commit()
        from app.utils.audit import log_action
        log_action(family_id=family_id, user_id=int(get_jwt_identity()),
                   action='excel_import', entity_type='member',
                   description=f'Excel 导入 {added} 个成员')
        return jsonify({
            'dry_run': False,
            'added': added,
            'invalid': len(errors),
            'errors': errors[:50],
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Excel 导入失败: {str(e)}'}), 500


@family_bp.route('/<int:family_id>/import/template', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def download_import_template(family_id):
    """下载导入模板 (.xlsx)"""
    import xlsxwriter
    from io import BytesIO

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('成员导入模板')

    header_format = wb.add_format({'bold': True, 'bg_color': '#b08d57', 'font_color': 'white',
                                    'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center'})
    example_format = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'vcenter'})

    headers = ['姓名', '性别', '出生日期', '逝世日期', '辈分', '字辈', '出生地', '逝世地',
               '是否在世', '字', '号', '谥', '简介']
    widths = [12, 8, 12, 12, 8, 10, 16, 16, 10, 10, 14, 10, 30]
    example = ['张三', '男', '1900-01-01', '1980-12-31', 1, '德', '山东济南', '北京',
               '否', '子明', '浩然居士', '文正', '家族第十二代传人，德高望重']

    for col, (h, w) in enumerate(zip(headers, widths)):
        ws.set_column(col, col, w)
        ws.write(0, col, h, header_format)

    for col, val in enumerate(example):
        ws.write(1, col, val, example_format)

    ws2 = wb.add_worksheet('字段说明')
    ws2.write(0, 0, '字段', header_format)
    ws2.write(0, 1, '说明', header_format)
    ws2.set_column(0, 0, 12)
    ws2.set_column(1, 1, 50)
    field_descs = [
        ('姓名', '必填，成员姓名'),
        ('性别', '可选：男 / 女 / 未知，默认未知'),
        ('出生日期', '可选，格式 YYYY-MM-DD 或 YYYY年MM月DD日'),
        ('逝世日期', '可选，格式同出生日期'),
        ('辈分', '可选，数字（第几代）'),
        ('字辈', '可选，该辈分名称（如：德、建、伟）'),
        ('出生地', '可选，出生地点'),
        ('逝世地', '可选，逝世地点'),
        ('是否在世', '可选，是/否/true/false，默认是'),
        ('字', '可选，古人的表字（如：子瞻）'),
        ('号', '可选，别号（如：东坡居士）'),
        ('谥', '可选，谥号（如：文忠）'),
        ('简介', '可选，生平简介文字'),
    ]
    for r, (f, d) in enumerate(field_descs):
        ws2.write(r + 1, 0, f, example_format)
        ws2.write(r + 1, 1, d, example_format)

    wb.close()
    output.seek(0)

    return output.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename=zupu_members_template.xlsx',
    }


# ==================== 关系管理 ====================

@family_bp.route('/<int:family_id>/relationships', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def list_relationships(family_id):
    """获取族谱所有关系"""
    relationships = FamilyService.get_family_relationships(family_id)
    return jsonify({
        'relationships': [r.to_dict() for r in relationships]
    })


@family_bp.route('/<int:family_id>/relationships', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def add_relationship(family_id):
    """添加成员关系"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    member_id = data.get('member_id')
    related_member_id = data.get('related_member_id')
    relationship_type = data.get('relationship_type')

    if not member_id or not related_member_id or not relationship_type:
        return jsonify({'error': '缺少必要参数'}), 400

    if relationship_type not in ('parent', 'spouse', 'sibling'):
        return jsonify({'error': '无效的关系类型'}), 400

    rel, error = FamilyService.add_relationship(
        member_id, related_member_id, relationship_type
    )
    if error:
        return jsonify({'error': error}), 400

    return jsonify({
        'message': '关系添加成功',
        'relationship': rel.to_dict()
    }), 201


@family_bp.route('/<int:family_id>/relationships/<int:relationship_id>', methods=['DELETE'])
@jwt_required()
@family_permission_required('editor')
def delete_relationship(family_id, relationship_id):
    """删除关系"""
    if FamilyService.remove_relationship(relationship_id):
        return jsonify({'message': '关系删除成功'})
    return jsonify({'error': '关系不存在'}), 404


# ==================== 族谱树形数据 ====================

@family_bp.route('/<int:family_id>/tree', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_family_tree(family_id):
    """获取族谱树形数据（用于可视化）"""
    tree_data = FamilyService.get_family_tree(family_id)
    role = _get_role(family_id)
    member_map = {}
    members = FamilyService.get_family_members(family_id)
    for m in members:
        member_map[m.id] = m
    from app.services.privacy_service import get_effective_privacy_level
    for node in tree_data.get('nodes', []):
        m = member_map.get(node['id'])
        if m:
            effective = get_effective_privacy_level(m)
            node['privacy_level'] = effective
            if effective == 'private' and role != 'admin':
                node['name'] = '***'
                node['gender'] = 'unknown'
                node.pop('birth_date', None)
                node.pop('death_date', None)
                node.pop('avatar', None)
                node['is_alive'] = None
    return jsonify(tree_data)


# ==================== 族谱统计 ====================

@family_bp.route('/<int:family_id>/stats', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_family_stats(family_id):
    """获取族谱统计数据"""
    from datetime import datetime
    members = FamilyService.get_family_members(family_id)
    relationships = FamilyService.get_family_relationships(family_id)

    total = len(members)
    male_count = sum(1 for m in members if m.gender == 'male')
    female_count = sum(1 for m in members if m.gender == 'female')
    unknown_count = sum(1 for m in members if m.gender == 'unknown')
    alive_count = sum(1 for m in members if m.is_alive)
    deceased_count = total - alive_count

    gen_stats = {}
    for m in members:
        if m.generation:
            gen_stats[f'第{m.generation}代'] = gen_stats.get(f'第{m.generation}代', 0) + 1

    parent_rels = sum(1 for r in relationships if r.relationship_type == 'parent')
    spouse_rels = sum(1 for r in relationships if r.relationship_type == 'spouse')
    sibling_rels = sum(1 for r in relationships if r.relationship_type == 'sibling')

    age_groups = {'0-17': {'male': 0, 'female': 0}, '18-35': {'male': 0, 'female': 0}, '36-60': {'male': 0, 'female': 0}, '60+': {'male': 0, 'female': 0}}
    current_year = datetime.now().year
    for m in members:
        if m.birth_date:
            try:
                birth_year = int(m.birth_date[:4])
                age = current_year - birth_year
                if age < 18:
                    age_key = '0-17'
                elif age < 36:
                    age_key = '18-35'
                elif age < 60:
                    age_key = '36-60'
                else:
                    age_key = '60+'
                if m.gender == 'male':
                    age_groups[age_key]['male'] += 1
                elif m.gender == 'female':
                    age_groups[age_key]['female'] += 1
            except:
                pass

    branch_stats: dict[int, dict] = {}
    for m in members:
        bid = getattr(m, 'branch_id', None) or 0
        bucket = branch_stats.setdefault(bid, {'count': 0, 'male': 0, 'female': 0, 'alive': 0})
        bucket['count'] += 1
        if m.gender == 'male':
            bucket['male'] += 1
        elif m.gender == 'female':
            bucket['female'] += 1
        if m.is_alive:
            bucket['alive'] += 1

    branch_labels = {}
    try:
        from app.models.branch import Branch
        for b in Branch.query.filter_by(family_id=family_id).all():
            branch_labels[b.id] = b.name
    except Exception:
        pass
    branch_distribution = []
    for bid, b in sorted(branch_stats.items(), key=lambda x: -x[1]['count']):
        branch_distribution.append({
            'branch_id': bid,
            'branch_name': branch_labels.get(bid, '未分配分支' if bid == 0 else f'分支#{bid}'),
            **b,
        })

    lifespans: list[dict] = []
    for m in members:
        if not m.birth_date:
            continue
        try:
            by = int(m.birth_date[:4])
            dy = None
            if m.death_date:
                dy = int(m.death_date[:4])
            elif not m.is_alive:
                continue
            if dy is not None:
                age = dy - by
                if 0 < age < 130:
                    lifespans.append({
                        'member_id': m.id,
                        'name': m.name,
                        'gender': m.gender,
                        'birth_year': by,
                        'death_year': dy,
                        'age': age,
                    })
        except (ValueError, TypeError):
            continue

    longevity = {
        'count': len(lifespans),
        'avg_age': round(sum(x['age'] for x in lifespans) / len(lifespans), 1) if lifespans else 0,
        'max_age': max((x['age'] for x in lifespans), default=0),
        'min_age': min((x['age'] for x in lifespans), default=0),
        'male_avg': round(sum(x['age'] for x in lifespans if x['gender'] == 'male') /
                          max(1, sum(1 for x in lifespans if x['gender'] == 'male')), 1),
        'female_avg': round(sum(x['age'] for x in lifespans if x['gender'] == 'female') /
                            max(1, sum(1 for x in lifespans if x['gender'] == 'female')), 1),
        'top_longest': sorted(lifespans, key=lambda x: -x['age'])[:5],
    }

    decade_distribution: dict[str, int] = {}
    for m in members:
        if not m.birth_date:
            continue
        try:
            by = int(m.birth_date[:4])
            decade = f"{(by // 10) * 10}s"
            decade_distribution[decade] = decade_distribution.get(decade, 0) + 1
        except (ValueError, TypeError):
            continue
    birth_decades = [{'decade': d, 'count': c} for d, c in sorted(decade_distribution.items())]

    return jsonify({
        'total': total,
        'gender': {
            'male': male_count,
            'female': female_count,
            'unknown': unknown_count,
        },
        'alive_status': {
            'alive': alive_count,
            'deceased': deceased_count,
        },
        'generation': gen_stats,
        'relationships': {
            'parent': parent_rels,
            'spouse': spouse_rels,
            'sibling': sibling_rels,
        },
        'age_groups': age_groups,
        'branch_distribution': branch_distribution,
        'longevity': longevity,
        'birth_decades': birth_decades,
    })


# ==================== 族谱用户权限管理 ====================

@family_bp.route('/<int:family_id>/users', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def list_family_users(family_id):
    """获取族谱的所有授权用户"""
    users = FamilyService.get_family_users(family_id)
    return jsonify({'users': users})


@family_bp.route('/<int:family_id>/users', methods=['POST'])
@jwt_required()
@family_permission_required('admin')
def add_family_user_route(family_id):
    """添加用户到族谱（授权）"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    user_id = data.get('user_id')
    role = data.get('role', 'viewer')

    if not user_id:
        return jsonify({'error': '用户ID不能为空'}), 400

    fm, error = FamilyService.add_family_user(family_id, user_id, role)
    if error:
        return jsonify({'error': error}), 400

    return jsonify({
        'message': '用户授权成功',
        'family_member': fm.to_dict()
    }), 201


@family_bp.route('/<int:family_id>/users/<int:user_id>', methods=['PUT'])
@jwt_required()
@family_permission_required('admin')
def update_family_user_role_route(family_id, user_id):
    """修改用户在族谱中的角色"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    new_role = data.get('role')
    if not new_role:
        return jsonify({'error': '角色不能为空'}), 400

    success, error = FamilyService.update_family_user_role(family_id, user_id, new_role)
    if not success:
        return jsonify({'error': error}), 400

    return jsonify({'message': '角色更新成功'})


@family_bp.route('/<int:family_id>/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@family_permission_required('admin')
def remove_family_user_route(family_id, user_id):
    """从族谱中移除用户"""
    success, error = FamilyService.remove_family_user(family_id, user_id)
    if not success:
        return jsonify({'error': error}), 400

    return jsonify({'message': '用户已移除'})


@family_bp.route('/users/search', methods=['GET'])
@jwt_required()
def search_users():
    """搜索用户（用于添加族谱成员时查找）"""
    query = request.args.get('q', '').strip()
    if not query or len(query) < 2:
        return jsonify({'error': '搜索关键词至少2个字符'}), 400

    users = User.query.filter(
        User.username.contains(query) | User.display_name.contains(query)
    ).limit(10).all()

    return jsonify({
        'users': [
            {
                'id': u.id,
                'username': u.username,
                'display_name': u.display_name,
                'avatar': u.avatar,
            }
            for u in users
        ]
    })


# ==================== 成员-用户账号绑定 ====================

@family_bp.route('/<int:family_id>/members/<int:member_id>/bound-user', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_bound_user(family_id, member_id):
    """获取成员绑定的用户账号"""
    member = Member.query.filter_by(id=member_id, family_id=family_id).first()
    if not member:
        return jsonify({'error': '成员不存在'}), 404

    bound_user = User.query.filter_by(linked_member_id=member_id).first()
    if not bound_user:
        return jsonify({'bound_user': None})

    return jsonify({
        'bound_user': {
            'id': bound_user.id,
            'username': bound_user.username,
            'display_name': bound_user.display_name,
            'avatar': bound_user.avatar,
            'email': bound_user.email if User.query.get(bound_user.id).id == int(get_jwt_identity()) else None,
        }
    })


@family_bp.route('/<int:family_id>/members/<int:member_id>/bound-user', methods=['PUT'])
@jwt_required()
@family_permission_required('admin')
def bind_user_to_member(family_id, member_id):
    """管理员将用户账号绑定到成员（替换原绑定）"""
    data = request.get_json() or {}
    user_id = data.get('user_id')

    if user_id is None:
        return jsonify({'error': '缺少 user_id'}), 400

    user_id = int(user_id)
    target_member = Member.query.filter_by(id=member_id, family_id=family_id).first()
    if not target_member:
        return jsonify({'error': '成员不存在'}), 404

    target_user = db.session.get(User, user_id)
    if not target_user:
        return jsonify({'error': '用户不存在'}), 404

    existing = User.query.filter(
        User.linked_member_id == member_id,
        User.id != user_id,
    ).first()
    if existing:
        existing.linked_member_id = None
        db.session.flush()

    old_binding = User.query.filter_by(linked_member_id=member_id).first()
    if old_binding and old_binding.id != user_id:
        old_binding.linked_member_id = None
        db.session.flush()

    target_user.linked_member_id = member_id
    db.session.commit()

    return jsonify({
        'message': '账号绑定成功',
        'bound_user': {
            'id': target_user.id,
            'username': target_user.username,
            'display_name': target_user.display_name,
            'avatar': target_user.avatar,
        },
    })


@family_bp.route('/<int:family_id>/members/<int:member_id>/bound-user', methods=['DELETE'])
@jwt_required()
@family_permission_required('admin')
def unbind_user_from_member(family_id, member_id):
    """解除成员的用户账号绑定"""
    target_member = Member.query.filter_by(id=member_id, family_id=family_id).first()
    if not target_member:
        return jsonify({'error': '成员不存在'}), 404

    bound = User.query.filter_by(linked_member_id=member_id).first()
    if not bound:
        return jsonify({'error': '该成员未绑定用户'}), 404

    bound.linked_member_id = None
    db.session.commit()
    return jsonify({'message': '已解除账号绑定'})


@family_bp.route('/<int:family_id>/my-linked-member', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def get_my_linked_member(family_id):
    """获取当前用户在族谱中绑定的成员"""
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user or not user.linked_member_id:
        return jsonify({'member': None})

    member = Member.query.filter_by(id=user.linked_member_id, family_id=family_id).first()
    if not member:
        return jsonify({'member': None})

    return jsonify({'member': member.to_dict()})


@family_bp.route('/<int:family_id>/my-linked-member', methods=['PUT'])
@jwt_required()
@family_permission_required('viewer')
def bind_self_to_member(family_id):
    """用户自己绑定到某成员（仅能绑定自己）"""
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    member_id = data.get('member_id')

    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    if member_id is None:
        user.linked_member_id = None
        db.session.commit()
        return jsonify({'message': '已解除自我关联'})

    member_id = int(member_id)
    target_member = Member.query.filter_by(id=member_id, family_id=family_id).first()
    if not target_member:
        return jsonify({'error': '成员不存在'}), 404

    exists = User.query.filter(
        User.linked_member_id == member_id,
        User.id != user_id,
    ).first()
    if exists:
        return jsonify({'error': '该成员已被其他账号绑定'}), 400

    user.linked_member_id = member_id
    db.session.commit()
    return jsonify({
        'message': '已将自己关联到此成员',
        'member': target_member.to_dict(),
    })


# ==================== 示例数据集 ====================

@family_bp.route('/sample-datasets', methods=['GET'])
@jwt_required()
def list_sample_datasets():
    """列出所有可用的内置示例数据集"""
    from app.seed_data import get_sample_dataset_list
    return jsonify({'datasets': get_sample_dataset_list()})


@family_bp.route('/<int:family_id>/load-sample', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def load_sample_data(family_id):
    """将示例数据加载到指定家族"""
    data = request.get_json(silent=True) or {}
    dataset_key = data.get('dataset_key')
    if not dataset_key:
        return jsonify({'error': '缺少 dataset_key 参数'}), 400

    from app.seed_data import load_sample_data as _load
    try:
        result = _load(family_id, dataset_key)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'加载失败: {str(e)}'}), 500

    from app.utils.audit import log_action
    log_action(
        family_id=family_id,
        user_id=int(get_jwt_identity()),
        action='load_sample',
        entity_type='family',
        description=f'加载示例数据集: {result["dataset_name"]}',
    )
    return jsonify(result)
