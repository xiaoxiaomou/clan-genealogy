from flask import Blueprint, request, jsonify, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.user import User
from app.models.generation_rule import GenerationRule
from app.models.member import Member
from app.models.family import Family, FamilyMember
from app.models.relationship import Relationship
from app.utils.decorators import family_permission_required, admin_required
import io

generation_bp = Blueprint('generation', __name__)


# ==================== 辈分字派管理 ====================

@generation_bp.route('/<int:family_id>/generations', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def list_generations(family_id):
    """获取族谱的辈分字派列表"""
    rules = GenerationRule.query.filter_by(
        family_id=family_id, is_active=True
    ).order_by(GenerationRule.generation.asc()).all()
    return jsonify({
        'generations': [r.to_dict() for r in rules]
    })


@generation_bp.route('/<int:family_id>/generations', methods=['POST'])
@jwt_required()
@family_permission_required('admin')
def add_generation(family_id):
    """添加辈分字派"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    generation = data.get('generation')
    character = data.get('character', '').strip()

    if not generation or not character:
        return jsonify({'error': '辈分和字派不能为空'}), 400

    # 检查是否已存在
    existing = GenerationRule.query.filter_by(
        family_id=family_id, generation=generation
    ).first()
    if existing:
        return jsonify({'error': f'第{generation}代的字派已存在'}), 400

    rule = GenerationRule(
        family_id=family_id,
        generation=generation,
        character=character,
        description=data.get('description', '').strip(),
    )
    db.session.add(rule)
    db.session.commit()
    return jsonify({
        'message': '辈分字派添加成功',
        'generation': rule.to_dict()
    }), 201


@generation_bp.route('/<int:family_id>/generations/<int:rule_id>', methods=['PUT'])
@jwt_required()
@family_permission_required('admin')
def update_generation(family_id, rule_id):
    """更新辈分字派"""
    rule = db.session.get(GenerationRule, rule_id)
    if not rule or rule.family_id != family_id:
        return jsonify({'error': '字派规则不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    if 'character' in data and data['character']:
        rule.character = data['character'].strip()
    if 'description' in data:
        rule.description = data.get('description', '').strip()
    if 'generation' in data and data['generation']:
        # 检查新的辈分是否冲突
        new_gen = data['generation']
        if new_gen != rule.generation:
            conflict = GenerationRule.query.filter_by(
                family_id=family_id, generation=new_gen
            ).first()
            if conflict:
                return jsonify({'error': f'第{new_gen}代的字派已存在'}), 400
            rule.generation = new_gen

    db.session.commit()
    return jsonify({
        'message': '辈分字派更新成功',
        'generation': rule.to_dict()
    })


@generation_bp.route('/<int:family_id>/generations/<int:rule_id>', methods=['DELETE'])
@jwt_required()
@family_permission_required('admin')
def delete_generation(family_id, rule_id):
    """删除辈分字派"""
    rule = db.session.get(GenerationRule, rule_id)
    if not rule or rule.family_id != family_id:
        return jsonify({'error': '字派规则不存在'}), 404

    rule.is_active = False  # 软删除
    db.session.commit()
    return jsonify({'message': '辈分字派已删除'})


@generation_bp.route('/<int:family_id>/generations/suggest', methods=['GET'])
@jwt_required()
@family_permission_required('viewer')
def suggest_generation(family_id):
    """为指定辈分推荐字派"""
    generation = request.args.get('generation', type=int)
    if not generation:
        return jsonify({'error': '请指定辈分'}), 400

    rule = GenerationRule.query.filter_by(
        family_id=family_id, generation=generation, is_active=True
    ).first()

    if rule:
        return jsonify({
            'has_rule': True,
            'character': rule.character,
            'description': rule.description,
        })
    else:
        # 返回最接近的辈分信息
        nearest = GenerationRule.query.filter_by(
            family_id=family_id, is_active=True
        ).order_by(
            db.func.abs(GenerationRule.generation - generation)
        ).first()
        return jsonify({
            'has_rule': False,
            'character': None,
            'nearest_generation': nearest.to_dict() if nearest else None,
        })


# ==================== 批量导入 ====================

@generation_bp.route('/<int:family_id>/import', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def import_members(family_id):
    """从 Excel/CSV 批量导入成员"""
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    filename = file.filename.lower()
    if filename.endswith('.csv'):
        return _import_csv(family_id, file)
    elif filename.endswith(('.xlsx', '.xls')):
        return _import_excel(family_id, file)
    else:
        return jsonify({'error': '仅支持 .csv 和 .xlsx/.xls 格式'}), 400


def _import_csv(family_id, file):
    """从 CSV 导入"""
    import csv
    import io

    stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
    reader = csv.DictReader(stream)

    required_fields = {'name'}
    if not reader.fieldnames:
        return jsonify({'error': 'CSV 文件格式不正确，无法读取表头'}), 400

    fields = set(f.strip() for f in reader.fieldnames)
    if not required_fields.issubset(fields):
        return jsonify({'error': f'CSV 文件缺少必填列: name（姓名）'}), 400

    added = []
    errors = []
    for row_idx, row in enumerate(reader, start=2):
        name = row.get('name', '').strip()
        if not name:
            errors.append(f'第{row_idx}行: 姓名为空')
            continue

        try:
            gen_val = row.get('generation', '').strip()
            generation = int(gen_val) if gen_val else None

            member = Member(
                family_id=family_id,
                name=name,
                gender=row.get('gender', 'unknown').strip().lower() or 'unknown',
                birth_date=row.get('birth_date', '').strip() or None,
                death_date=row.get('death_date', '').strip() or None,
                generation=generation,
                generation_name=row.get('generation_name', '').strip() or None,
                bio=row.get('bio', '').strip() or None,
                is_alive=row.get('is_alive', '1').strip() in ('1', 'yes', 'true', '是'),
            )
            db.session.add(member)
            db.session.flush()
            added.append(member)
        except Exception as e:
            errors.append(f'第{row_idx}行 ({name}): {str(e)}')

    db.session.commit()

    return jsonify({
        'message': f'成功导入 {len(added)} 名成员',
        'added_count': len(added),
        'error_count': len(errors),
        'errors': errors[:20],  # 最多返回20条错误
    })


def _import_excel(family_id, file):
    """从 Excel 导入"""
    import openpyxl

    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    if not ws:
        return jsonify({'error': 'Excel 文件没有工作表'}), 400

    # 读取表头
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip().lower() if cell else '')

    if 'name' not in headers:
        return jsonify({'error': 'Excel 文件缺少 "name"（姓名）列'}), 400

    col_map = {h: i for i, h in enumerate(headers)}
    added = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_vals = [str(v).strip() if v else '' for v in row]
        if not any(row_vals):
            continue  # 跳过空行

        name = row_vals[col_map.get('name', 0)] if len(row_vals) > col_map.get('name', 0) else ''
        if not name:
            errors.append(f'第{row_idx}行: 姓名为空')
            continue

        try:
            get_val = lambda key: row_vals[col_map[key]] if key in col_map and len(row_vals) > col_map[key] else ''

            gen_val = get_val('generation')
            generation = int(gen_val) if gen_val else None

            member = Member(
                family_id=family_id,
                name=name,
                gender=get_val('gender').lower() or 'unknown',
                birth_date=get_val('birth_date') or None,
                death_date=get_val('death_date') or None,
                generation=generation,
                generation_name=get_val('generation_name') or None,
                bio=get_val('bio') or None,
                is_alive=get_val('is_alive') in ('1', 'yes', 'true', '是'),
            )
            db.session.add(member)
            db.session.flush()
            added.append(member)
        except Exception as e:
            errors.append(f'第{row_idx}行 ({name}): {str(e)}')

    db.session.commit()
    wb.close()

    return jsonify({
        'message': f'成功导入 {len(added)} 名成员',
        'added_count': len(added),
        'error_count': len(errors),
        'errors': errors[:20],
    })


# ==================== 快速建家庭 ====================

@generation_bp.route('/<int:family_id>/quick-family', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def quick_add_family(family_id):
    """快速添加一个家庭（父亲 + 母亲 + 子女列表）"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求数据不能为空'}), 400

    father_data = data.get('father', {})
    mother_data = data.get('mother')
    children_data = data.get('children', [])

    father_name = father_data.get('name', '').strip()
    if not father_name:
        return jsonify({'error': '父亲姓名不能为空'}), 400

    created_members = []

    # 1. 创建父亲
    father = Member(
        family_id=family_id,
        name=father_name,
        gender='male',
        birth_date=father_data.get('birth_date') or None,
        generation=father_data.get('generation'),
        generation_name=father_data.get('generation_name'),
        bio=father_data.get('bio'),
        is_alive=father_data.get('is_alive', True),
    )
    db.session.add(father)
    db.session.flush()
    created_members.append({'member': father, 'role': 'father'})

    # 2. 创建母亲（可选）
    mother = None
    if mother_data and mother_data.get('name', '').strip():
        mother = Member(
            family_id=family_id,
            name=mother_data['name'].strip(),
            gender='female',
            birth_date=mother_data.get('birth_date') or None,
            generation=mother_data.get('generation'),
            generation_name=mother_data.get('generation_name'),
            bio=mother_data.get('bio'),
            is_alive=mother_data.get('is_alive', True),
        )
        db.session.add(mother)
        db.session.flush()
        created_members.append({'member': mother, 'role': 'mother'})

        # 建立配偶关系（双向）
        rel = Relationship(
            member_id=father.id,
            related_member_id=mother.id,
            relationship_type='spouse'
        )
        db.session.add(rel)
        rel2 = Relationship(
            member_id=mother.id,
            related_member_id=father.id,
            relationship_type='spouse'
        )
        db.session.add(rel2)

    # 3. 创建子女
    children_created = []
    for idx, child in enumerate(children_data):
        child_name = child.get('name', '').strip()
        if not child_name:
            continue

        child_gen = child.get('generation') or (
            (father.generation + 1) if father.generation else None
        )

        member = Member(
            family_id=family_id,
            name=child_name,
            gender=child.get('gender', 'unknown'),
            birth_date=child.get('birth_date') or None,
            generation=child_gen,
            generation_name=None,
            is_alive=child.get('is_alive', True),
        )
        db.session.add(member)
        db.session.flush()
        children_created.append(member)
        created_members.append({'member': member, 'role': 'child'})

        # 建立父子关系
        rel_parent = Relationship(
            member_id=father.id,
            related_member_id=member.id,
            relationship_type='parent'
        )
        db.session.add(rel_parent)

        # 如果有母亲，也建立母子关系
        if mother:
            rel_mother = Relationship(
                member_id=mother.id,
                related_member_id=member.id,
                relationship_type='parent'
            )
            db.session.add(rel_mother)

    # 4. 兄弟姊妹关系
    for i in range(len(children_created)):
        for j in range(i + 1, len(children_created)):
            rel_sibling = Relationship(
                member_id=children_created[i].id,
                related_member_id=children_created[j].id,
                relationship_type='sibling'
            )
            db.session.add(rel_sibling)

    db.session.commit()

    return jsonify({
        'message': f'家庭创建成功：父亲 {father_name}' +
                   (f'、母亲 {mother_data["name"]}' if mother else '') +
                   f'、{len(children_created)} 名子女',
        'father': father.to_dict(),
        'mother': mother.to_dict() if mother else None,
        'children': [c.to_dict() for c in children_created],
    }), 201


def _create_member(family_id, data, default_gender='unknown'):
    """创建单个成员"""
    if not data or not data.get('name', '').strip():
        return None
    return Member(
        family_id=family_id,
        name=data['name'].strip(),
        gender=data.get('gender', default_gender),
        birth_date=data.get('birth_date') or None,
        death_date=data.get('death_date') or None,
        generation=data.get('generation'),
        generation_name=data.get('generation_name'),
        bio=data.get('bio'),
        is_alive=data.get('is_alive', True),
    )


def _add_relationship(member_id, related_id, rel_type):
    """添加单向关系"""
    rel = Relationship(member_id=member_id, related_member_id=related_id, relationship_type=rel_type)
    db.session.add(rel)


def _link_spouses(m1, m2):
    """建立双向配偶关系"""
    _add_relationship(m1.id, m2.id, 'spouse')
    _add_relationship(m2.id, m1.id, 'spouse')


def _link_parent(parent, child):
    """建立亲子关系"""
    _add_relationship(parent.id, child.id, 'parent')


def _link_siblings(children):
    """建立兄弟姐妹关系"""
    for i in range(len(children)):
        for j in range(i + 1, len(children)):
            _add_relationship(children[i].id, children[j].id, 'sibling')


@generation_bp.route('/<int:family_id>/quick-family-multi', methods=['POST'])
@jwt_required()
@family_permission_required('editor')
def quick_add_multi_family(family_id):
    """快速创建多代家庭"""
    data = request.get_json()
    if not data or 'layers' not in data:
        return jsonify({'error': '缺少 layers 参数'}), 400

    layers = data['layers']
    if not layers:
        return jsonify({'error': '至少需要一代'}), 400

    all_created = []
    # 记录每层已创建的成员名 → Member，用于跨代匹配
    prev_layer_members = {}  # name -> Member

    for layer_idx, layer in enumerate(layers):
        father_data = layer.get('father', {})
        mother_data = layer.get('mother')
        children_data = layer.get('children', [])

        father_name = father_data.get('name', '').strip()
        if not father_name:
            return jsonify({'error': f'第 {layer_idx + 1} 代缺少父亲姓名'}), 400

        # 如果父亲姓名在上一层的成员中已存在，复用该成员
        father = prev_layer_members.get(father_name)
        if not father:
            father = _create_member(family_id, father_data, 'male')
            if not father:
                return jsonify({'error': f'第 {layer_idx + 1} 代父亲数据无效'}), 400
            db.session.add(father)
            db.session.flush()
            all_created.append({'member': father, 'role': f'layer{layer_idx}_father'})

        mother = None
        if mother_data and mother_data.get('name', '').strip():
            mother = _create_member(family_id, mother_data, 'female')
            if mother:
                db.session.add(mother)
                db.session.flush()
                all_created.append({'member': mother, 'role': f'layer{layer_idx}_mother'})
                _link_spouses(father, mother)

        # 创建子女
        children_created = []
        for child in children_data:
            child_name = child.get('name', '').strip()
            if not child_name:
                continue

            child_gen = child.get('generation') or (
                (father.generation + 1) if father.generation else None
            )
            child_obj = Member(
                family_id=family_id,
                name=child_name,
                gender=child.get('gender', 'unknown'),
                birth_date=child.get('birth_date') or None,
                generation=child_gen,
                generation_name=None,
                is_alive=child.get('is_alive', True),
            )
            db.session.add(child_obj)
            db.session.flush()
            children_created.append(child_obj)
            all_created.append({'member': child_obj, 'role': f'layer{layer_idx}_child'})

            _link_parent(father, child_obj)
            if mother:
                _link_parent(mother, child_obj)

            # 记录到跨代映射表
            if child_name not in prev_layer_members:
                prev_layer_members[child_name] = child_obj

        _link_siblings(children_created)

        # 如果父亲是新建的（非复用），也加入映射表
        if father_name not in prev_layer_members:
            prev_layer_members[father_name] = father

    db.session.commit()

    # 按层组织返回数据
    result_layers = []
    for layer_idx, layer in enumerate(layers):
        layer_data = {'father': None, 'mother': None, 'children': []}
        for entry in all_created:
            role = entry['role']
            if role == f'layer{layer_idx}_father':
                layer_data['father'] = entry['member'].to_dict()
            elif role == f'layer{layer_idx}_mother':
                layer_data['mother'] = entry['member'].to_dict()
            elif role == f'layer{layer_idx}_child':
                layer_data['children'].append(entry['member'].to_dict())
        result_layers.append(layer_data)

    total = len(all_created)
    return jsonify({
        'message': f'多代家庭创建成功，共 {total} 人，{len(layers)} 代',
        'layers': result_layers,
    }), 201


@generation_bp.route('/<int:family_id>/template/excel', methods=['GET'])
@jwt_required()
@family_permission_required('editor')
def download_template_excel(family_id):
    """下载成员导入模板 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成员导入模板'

    headers = ['name', 'gender', 'birth_date', 'death_date', 'generation', 'generation_name', 'bio', 'is_alive']
    header_labels = ['姓名*', '性别(male/female/unknown)', '出生日期(YYYY-MM-DD)', '逝世日期(YYYY-MM-DD)', '辈分(数字)', '字辈', '简介', '是否在世(1/0)']

    header_fill = PatternFill(start_color='8b2500', end_color='8b2500', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, (h, label) in enumerate(zip(headers, header_labels), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    example_data = [
        ['张三', 'male', '1960-05-01', '', 3, '德', '家族长者', 1],
        ['李四', 'female', '1965-03-15', '', 3, '贤', '家族贤内助', 1],
        ['王小明', 'male', '1990-08-20', '', 4, '文', '清华大学毕业', 1],
        ['张老六', 'male', '1930-01-01', '2020-12-31', 2, '福', '已故', 0],
    ]

    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=member_import_template.xlsx'
    return response


@generation_bp.route('/<int:family_id>/template/csv', methods=['GET'])
@jwt_required()
@family_permission_required('editor')
def download_template_csv(family_id):
    """下载成员导入模板 CSV"""
    import csv

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ['name', 'gender', 'birth_date', 'death_date', 'generation', 'generation_name', 'bio', 'is_alive']
    header_labels = ['姓名*', '性别(male/female/unknown)', '出生日期(YYYY-MM-DD)', '逝世日期(YYYY-MM-DD)', '辈分(数字)', '字辈', '简介', '是否在世(1/0)']
    writer.writerow(header_labels)

    example_data = [
        ['张三', 'male', '1960-05-01', '', 3, '德', '家族长者', 1],
        ['李四', 'female', '1965-03-15', '', 3, '贤', '家族贤内助', 1],
        ['王小明', 'male', '1990-08-20', '', 4, '文', '清华大学毕业', 1],
        ['张老六', 'male', '1930-01-01', '2020-12-31', 2, '福', '已故', 0],
    ]
    for row in example_data:
        writer.writerow(row)

    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv;charset=utf-8-sig'
    response.headers['Content-Disposition'] = 'attachment; filename=member_import_template.csv'
    return response
